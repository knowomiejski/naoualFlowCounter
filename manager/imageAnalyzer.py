

import scyjava as sj
from openpyxl import load_workbook
from openpyxl import Workbook
from os import remove
from os import popen
from os import mkdir
from os import listdir
from os import fsdecode
from os import path
import time
import imagej
import glob
import csv


class ImageAnalyzer:

    def __init__(self):
        self.ij = imagej.init("imagej/Fiji.app", mode="interactive")
        # self.ij = imagej.init("imagej/Fiji.segmentation.app", mode="interactive")
        # self.ij.IJ.run("Console", "")
        # self.ij.IJ.run("Record...", "");
        print("STARTING IMAGEJ...")
        print(self.ij.getVersion())

    def countFlowCells(self, fileA, fileB):
        print("FileA:", fileA);
        fileAName = path.basename(fileA)
        fileBName = path.basename(fileB)
        print("FileB:", fileB);
        fileAData = self.ij.IJ.openImage(fileA)
        fileBData = self.ij.IJ.openImage(fileB)
        self.ij.ui().show(fileAData)
        self.ij.IJ.run("FlowImageCounting pt1", "");
        self.ij.ui().show(fileBData)
        self.ij.IJ.run("FlowImageCounting pt1", "");


        imp = self.ij.py.active_imageplus()

        self.ij.IJ.run(imp, "FlowImageCounting pt2", "image1=["+fileAName+"] operation=Multiply image2=["+fileBName+"] create");
        results = self.ij.ResultsTable.getResultsTable()
        csvfilePath = path.join("data", "outputs", fileAName.split(".")[0] + ".csv")
        results.saveAs(csvfilePath)
        if (path.exists(path.join("data", "outputs", "Results.xlsx")) == False):
            wb = Workbook()
            ws = wb.worksheets[0]
            titles = [
                "Image",
                "Count",
                "Area",
                "Perim",
                "Major",
                "Minor",
                "Angle",
                "Circ",
                "AR",
                "Round",
                "Solidity"
            ]
            ws.append(titles)
            wb.save(path.join("data", "outputs", "Results.xlsx"))

        csvfile = open(csvfilePath, "r")
        reader = csv.reader(csvfile, delimiter=",")
        data = list(reader)
        lastRowRaw = data[-1]
        lastRowRaw.insert(0, fileAName + "/" + fileBName)
        lastRow = []
        for row in lastRowRaw:
            if self.is_float(row):
                lastRow.append(float(row))
            else:
                lastRow.append(row)
        print("lastRow", lastRow)

        wb = load_workbook(path.join("data", "outputs", "Results.xlsx"))
        ws = wb.worksheets[0]
        ws.append(lastRow)
        wb.save(path.join("data", "outputs", "Results.xlsx"))

        self.ij.IJ.run("Close All", "");
        self.ij.IJ.selectWindow("Results");
        self.ij.IJ.run("Close");
        roiManager = self.ij.RoiManager.getRoiManager();
        roiManager.close();
        csvfile.close()
        # remove(csvfilePath)

    def is_float(self, string):
        if string.replace(".", "").isnumeric():
            return True
        else:
            return False
